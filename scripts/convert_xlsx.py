#!/usr/bin/env python3
"""Convert an English vocabulary workbook into the app's JSON schema."""

from __future__ import annotations

import argparse
import json
import re
import sys
import zipfile
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


ALIASES = {
    "word": {
        "word",
        "words",
        "vocab",
        "vocabulary",
        "english",
        "eng",
        "term",
        "단어",
        "영단어",
        "어휘",
        "표제어",
    },
    "meaningKo": {
        "meaning",
        "meanings",
        "definition",
        "definitions",
        "korean",
        "kor",
        "translation",
        "뜻",
        "의미",
        "해석",
        "우리말",
        "한국어",
    },
    "partOfSpeech": {
        "partofspeech",
        "part of speech",
        "pos",
        "품사",
    },
    "exampleEn": {
        "example",
        "exampleen",
        "example sentence",
        "sentence",
        "예문",
        "영어예문",
    },
    "exampleKo": {
        "exampleko",
        "example translation",
        "sentence meaning",
        "예문해석",
        "예문뜻",
    },
    "unit": {
        "unit",
        "day",
        "chapter",
        "lesson",
        "section",
        "단원",
        "챕터",
        "강",
        "일차",
    },
    "level": {
        "level",
        "difficulty",
        "rank",
        "grade",
        "난이도",
        "등급",
    },
    "tags": {
        "tag",
        "tags",
        "category",
        "categories",
        "분류",
        "태그",
    },
}

CANONICAL_FIELDS = tuple(ALIASES.keys())
XML_NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "pkgrel": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ").strip()
    text = re.sub(r"[ \t]+", " ", text)
    return text


def normalize_header(value: Any) -> str:
    text = clean_text(value).lower()
    text = re.sub(r"[\s_\-:/()[\]{}]+", " ", text).strip()
    squashed = re.sub(r"\s+", "", text)
    return squashed or text


def canonical_for_header(value: Any) -> str | None:
    raw = clean_text(value).lower()
    normalized = normalize_header(value)
    for field, aliases in ALIASES.items():
        for alias in aliases:
            alias_raw = alias.lower()
            alias_normalized = normalize_header(alias)
            if raw == alias_raw or normalized == alias_normalized:
                return field
    return None


def split_list(text: str) -> list[str]:
    if not text:
        return []
    chunks = re.split(r"\s*(?:\n|;|/|,|·|ㆍ)\s*", text)
    return [chunk for chunk in (clean_text(part) for part in chunks) if chunk]


def slugify(text: str) -> str:
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", text.lower()).strip("-")
    return slug or "word"


def looks_like_index(text: str) -> bool:
    return bool(re.fullmatch(r"\d+(?:\.0)?", clean_text(text)))


def normalize_unit_label(label: str, sheet_name: str, group_index: int) -> str:
    label = clean_text(label)
    sheet_numbers = [int(value) for value in re.findall(r"\d+", sheet_name)]
    sheet_unit = ""
    if 1 <= group_index <= len(sheet_numbers):
        sheet_unit = f"DAY {sheet_numbers[group_index - 1]:02d}"

    if re.fullmatch(r"DAY\s+\d{1,3}", label, flags=re.IGNORECASE):
        day_number = int(re.search(r"\d+", label).group(0))
        if sheet_unit and day_number != sheet_numbers[group_index - 1]:
            return sheet_unit
        return f"DAY {day_number:02d}"

    if sheet_unit:
        return sheet_unit
    return label or sheet_name


def row_values(row: tuple[Any, ...]) -> list[str]:
    return [clean_text(cell) for cell in row]


def column_index(cell_ref: str) -> int:
    letters = re.sub(r"[^A-Z]", "", cell_ref.upper())
    index = 0
    for char in letters:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return max(index - 1, 0)


def xml_text(element: ET.Element | None) -> str:
    if element is None:
        return ""
    return "".join(element.itertext())


def read_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    try:
        root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    except KeyError:
        return []
    strings = []
    for item in root.findall("main:si", XML_NS):
        strings.append(clean_text(xml_text(item)))
    return strings


def workbook_sheet_paths(archive: zipfile.ZipFile) -> list[tuple[str, str]]:
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    rel_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rels = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rel_root.findall("pkgrel:Relationship", XML_NS)
        if "Id" in rel.attrib and "Target" in rel.attrib
    }
    sheets = []
    for sheet in workbook_root.findall("main:sheets/main:sheet", XML_NS):
        rel_id = sheet.attrib.get(f"{{{XML_NS['rel']}}}id")
        target = rels.get(rel_id or "")
        if not target:
            continue
        path = target if target.startswith("xl/") else f"xl/{target.lstrip('/')}"
        sheets.append((sheet.attrib.get("name", "Sheet"), path))
    return sheets


def cell_value(cell: ET.Element, shared_strings: list[str]) -> str:
    cell_type = cell.attrib.get("t", "")
    if cell_type == "inlineStr":
        return clean_text(xml_text(cell.find("main:is", XML_NS)))

    value_node = cell.find("main:v", XML_NS)
    if value_node is None:
        return ""
    value = clean_text(value_node.text)
    if cell_type == "s":
        try:
            return shared_strings[int(value)]
        except (ValueError, IndexError):
            return value
    if cell_type == "b":
        return "TRUE" if value == "1" else "FALSE"
    return value


def parse_sheet_rows(archive: zipfile.ZipFile, sheet_path: str, shared_strings: list[str]) -> list[list[str]]:
    root = ET.fromstring(archive.read(sheet_path))
    rows: list[list[str]] = []
    for row in root.findall("main:sheetData/main:row", XML_NS):
        values: dict[int, str] = {}
        for cell in row.findall("main:c", XML_NS):
            cell_ref = cell.attrib.get("r", "")
            index = column_index(cell_ref) if cell_ref else len(values)
            values[index] = cell_value(cell, shared_strings)
        if values:
            width = max(values) + 1
            rows.append([values.get(index, "") for index in range(width)])
        else:
            rows.append([])
    return rows


def make_record(
    word: str,
    meaning: str,
    unit: str,
    sheet_name: str,
    row_number: int,
    seen_ids: defaultdict[str, int],
    raw: dict[str, str] | None = None,
) -> dict[str, Any] | None:
    word = clean_text(word)
    meaning = clean_text(meaning)
    if not word or not meaning:
        return None

    base_id = slugify(word)
    seen_ids[base_id] += 1
    suffix = "" if seen_ids[base_id] == 1 else f"-{seen_ids[base_id]}"
    word_id = f"{base_id}{suffix}"

    return {
        "id": word_id,
        "word": word,
        "meaningKo": meaning,
        "meanings": split_list(meaning) or [meaning],
        "partOfSpeech": "",
        "exampleEn": "",
        "exampleKo": "",
        "unit": clean_text(unit) or sheet_name,
        "level": "",
        "tags": [],
        "source": {
            "sheet": sheet_name,
            "row": row_number,
        },
        "raw": raw or {},
    }


def paired_column_groups(rows: list[list[str]]) -> list[tuple[int, int, int]]:
    if not rows:
        return []
    max_width = max(len(row) for row in rows)
    groups: list[tuple[int, int, int]] = []
    for start in range(max_width - 2):
        score = 0
        for row in rows[1:15]:
            if len(row) <= start + 2:
                continue
            if looks_like_index(row[start]) and clean_text(row[start + 1]) and clean_text(row[start + 2]):
                score += 1
        if score >= 3:
            groups.append((start, start + 1, start + 2))
    return groups


def convert_paired_rows(
    sheet_name: str,
    rows: list[list[str]],
    seen_ids: defaultdict[str, int],
) -> tuple[list[dict[str, Any]], dict[str, str]] | None:
    groups = paired_column_groups(rows)
    if len(groups) < 2:
        return None

    records: list[dict[str, Any]] = []
    detected: dict[str, str] = {}
    header = rows[0] if rows else []

    for group_index, (index_col, word_col, meaning_col) in enumerate(groups, start=1):
        header_label = header[index_col] if index_col < len(header) else ""
        unit = normalize_unit_label(header_label, sheet_name, group_index)
        detected[f"group_{group_index}_index"] = f"Column {index_col + 1}"
        detected[f"group_{group_index}_word"] = f"Column {word_col + 1}"
        detected[f"group_{group_index}_meaningKo"] = f"Column {meaning_col + 1}"

        for row_number, values in enumerate(rows[1:], start=2):
            if len(values) <= meaning_col:
                continue
            if not looks_like_index(values[index_col]):
                continue
            record = make_record(
                values[word_col],
                values[meaning_col],
                unit,
                sheet_name,
                row_number,
                seen_ids,
                raw={
                    "index": clean_text(values[index_col]),
                    "column_group": str(group_index),
                },
            )
            if record:
                records.append(record)

    return records, detected


def convert_rows(
    sheet_rows: list[tuple[str, list[list[str]]]],
    source: str,
    parser: str,
) -> dict[str, Any]:
    records: list[dict[str, Any]] = []
    seen_ids: defaultdict[str, int] = defaultdict(int)
    detected_columns: dict[str, dict[str, str]] = {}

    for sheet_name, rows in sheet_rows:
        rows = [values for values in rows if any(values)]
        if not rows:
            continue

        paired = convert_paired_rows(sheet_name, rows, seen_ids)
        if paired is not None:
            paired_records, detected = paired
            records.extend(paired_records)
            detected_columns[sheet_name] = detected
            continue

        header_index, mapping = detect_header_row(rows)
        start_index = 0
        if header_index is not None:
            start_index = header_index + 1
            header_values = rows[header_index]
        else:
            mapping = infer_columns(rows[0])
            header_values = [f"Column {index + 1}" for index in range(len(rows[0]))]

        detected_columns[sheet_name] = {
            field: header_values[index] if index < len(header_values) else f"Column {index + 1}"
            for index, field in mapping.items()
        }

        for idx, values in enumerate(rows[start_index:], start=start_index + 1):
            record = record_from_row(values, mapping, sheet_name, idx, seen_ids)
            if record:
                records.append(record)

    return {
        "metadata": {
            "schemaVersion": 1,
            "title": Path(source).stem,
            "source": source,
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "wordCount": len(records),
            "parser": parser,
            "detectedColumns": detected_columns,
        },
        "words": records,
    }


def detect_header_row(rows: list[list[str]]) -> tuple[int | None, dict[int, str]]:
    best_index: int | None = None
    best_map: dict[int, str] = {}
    best_score = 0

    for idx, values in enumerate(rows[:20]):
        current: dict[int, str] = {}
        for col_index, value in enumerate(values):
            field = canonical_for_header(value)
            if field and field not in current.values():
                current[col_index] = field
        score = len(set(current.values()))
        if "word" in current.values():
            score += 2
        if "meaningKo" in current.values():
            score += 2
        if score > best_score:
            best_index = idx
            best_map = current
            best_score = score

    if best_score >= 4:
        return best_index, best_map
    return None, {}


def infer_columns(values: list[str]) -> dict[int, str]:
    non_empty = [index for index, value in enumerate(values) if value]
    mapping: dict[int, str] = {}
    if non_empty:
        mapping[non_empty[0]] = "word"
    if len(non_empty) > 1:
        mapping[non_empty[1]] = "meaningKo"
    if len(non_empty) > 2:
        mapping[non_empty[2]] = "partOfSpeech"
    if len(non_empty) > 3:
        mapping[non_empty[3]] = "exampleEn"
    return mapping


def record_from_row(
    values: list[str],
    mapping: dict[int, str],
    sheet_name: str,
    row_number: int,
    seen_ids: defaultdict[str, int],
) -> dict[str, Any] | None:
    data = {field: "" for field in CANONICAL_FIELDS}
    raw: dict[str, str] = {}

    for col_index, value in enumerate(values):
        if not value:
            continue
        field = mapping.get(col_index)
        if field:
            data[field] = value
        else:
            raw[f"column_{col_index + 1}"] = value

    word = clean_text(data["word"])
    meaning = clean_text(data["meaningKo"])
    if not word or not meaning:
        return None

    base_id = slugify(word)
    seen_ids[base_id] += 1
    suffix = "" if seen_ids[base_id] == 1 else f"-{seen_ids[base_id]}"
    word_id = f"{base_id}{suffix}"

    tags = split_list(data["tags"])
    unit = data["unit"] or sheet_name

    return {
        "id": word_id,
        "word": word,
        "meaningKo": meaning,
        "meanings": split_list(meaning) or [meaning],
        "partOfSpeech": clean_text(data["partOfSpeech"]),
        "exampleEn": clean_text(data["exampleEn"]),
        "exampleKo": clean_text(data["exampleKo"]),
        "unit": clean_text(unit),
        "level": clean_text(data["level"]),
        "tags": tags,
        "source": {
            "sheet": sheet_name,
            "row": row_number,
        },
        "raw": raw,
    }


def convert_workbook(input_path: Path) -> dict[str, Any]:
    try:
        workbook = load_workbook(input_path, read_only=True, data_only=True)
        sheet_rows = [
            (sheet.title, [row_values(row) for row in sheet.iter_rows(values_only=True)])
            for sheet in workbook.worksheets
        ]
        return convert_rows(sheet_rows, str(input_path), "openpyxl")
    except Exception as exc:
        print(f"openpyxl failed, falling back to XML parser: {exc}", file=sys.stderr)

    with zipfile.ZipFile(input_path) as archive:
        shared_strings = read_shared_strings(archive)
        sheet_rows = [
            (sheet_name, parse_sheet_rows(archive, sheet_path, shared_strings))
            for sheet_name, sheet_path in workbook_sheet_paths(archive)
        ]
    return convert_rows(sheet_rows, str(input_path), "xlsx-xml")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path, help="Path to the source .xlsx file")
    parser.add_argument(
        "output",
        type=Path,
        nargs="?",
        default=Path("data/words.json"),
        help="Output JSON path. Defaults to data/words.json",
    )
    args = parser.parse_args()

    if not args.input.exists():
        print(f"Input file not found: {args.input}", file=sys.stderr)
        return 1
    if args.input.suffix.lower() not in {".xlsx", ".xlsm"}:
        print("Input must be an .xlsx or .xlsm workbook.", file=sys.stderr)
        return 1

    payload = convert_workbook(args.input)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"Wrote {len(payload['words'])} words to {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
